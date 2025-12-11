#!/usr/bin/env python3
"""
CSV to XLSX Converter for Pascal Legacy Telemetry
Converts typed CSV files to Excel format with proper data types
"""

import os
import sys
import glob
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("[ERROR] Required packages not installed. Installing...")
    os.system("pip3 install pandas openpyxl --quiet")
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows


def convert_csv_to_xlsx(csv_path: str) -> str:
    """Convert CSV to XLSX with proper data types and formatting"""
    try:
        # Read CSV
        df = pd.read_csv(csv_path)
        
        # Parse data types
        if 'Timestamp' in df.columns:
            df['Timestamp'] = pd.to_datetime(df['Timestamp'], errors='coerce')
        
        if 'UnixTimestamp' in df.columns:
            df['UnixTimestamp'] = pd.to_numeric(df['UnixTimestamp'], errors='coerce').astype('Int64')
        
        if 'Voltage' in df.columns:
            df['Voltage'] = pd.to_numeric(df['Voltage'], errors='coerce')
        
        if 'Temperature' in df.columns:
            df['Temperature'] = pd.to_numeric(df['Temperature'], errors='coerce')
        
        if 'IsActive' in df.columns:
            df['IsActive'] = df['IsActive'].map({'TRUE': True, 'FALSE': False, True: True, False: False})
        
        # Create XLSX
        xlsx_path = csv_path.replace('.csv', '.xlsx')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Telemetry Data"
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header styling
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Data formatting
                else:
                    if c_idx == 1 and isinstance(value, datetime):  # Timestamp
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    elif c_idx == 2:  # UnixTimestamp
                        cell.number_format = '0'
                    elif c_idx in [3, 4]:  # Voltage, Temperature
                        cell.number_format = '0.00'
                    elif c_idx == 5:  # IsActive
                        cell.value = 'TRUE' if value else 'FALSE'
                        cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(xlsx_path)
        
        print(f"[INFO] Converted: {csv_path} -> {xlsx_path}")
        return xlsx_path
        
    except Exception as e:
        print(f"[ERROR] Failed to convert {csv_path}: {e}")
        return None


def main():
    csv_dir = os.getenv('CSV_OUT_DIR', '/data/csv')
    
    if not os.path.exists(csv_dir):
        print(f"[ERROR] CSV directory not found: {csv_dir}")
        sys.exit(1)
    
    print(f"[INFO] Scanning for CSV files in: {csv_dir}")
    
    csv_files = glob.glob(f"{csv_dir}/telemetry_*.csv")
    
    if not csv_files:
        print("[WARN] No CSV files found")
        return
    
    converted_count = 0
    for csv_file in csv_files:
        xlsx_file = csv_file.replace('.csv', '.xlsx')
        
        # Convert if XLSX doesn't exist or CSV is newer
        if not os.path.exists(xlsx_file) or os.path.getmtime(csv_file) > os.path.getmtime(xlsx_file):
            if convert_csv_to_xlsx(csv_file):
                converted_count += 1
    
    print(f"[INFO] Conversion complete: {converted_count} file(s) processed")


if __name__ == "__main__":
    main()
